"""
build_nztcs_keys.py
===================
Augments nztcs.json with GBIF taxon-key based entries ("gbifkey:XXXXXX").

For each NZTCS species name, calls the GBIF Species Match API to get the
authoritative taxon key(s). Adds entries keyed by "gbifkey:XXXXXX" to
nztcs.json so that lookups using GBIF record taxonKey are exact and
immune to name variations (genus reclassifications, synonyms, etc.).

Run AFTER build_nztcs.py. Safe to re-run — uses a local cache so only
new/uncached names hit the API.

Usage:
    python build_nztcs_keys.py
"""

import sys, json, re, time
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from urllib.request import urlopen
from urllib.parse import urlencode
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter

XLSX       = r'C:\Users\User\Downloads\NZTCS Exported Data.xlsx'
NZTCS_JSON = r'C:\Users\User\GBIF-Record-Finder\.claude\worktrees\infallible-babbage-f27395\nztcs.json'
CACHE_FILE = r'C:\Users\User\GBIF-Record-Finder\.claude\worktrees\infallible-babbage-f27395\_gbif_match_cache.json'

CONCURRENCY  = 20      # parallel requests
CONFIDENCE   = 85      # minimum GBIF match confidence to accept
MATCH_TYPES  = {'EXACT', 'FUZZY'}

# ── Helpers ────────────────────────────────────────────────────────────────

def clean(s):
    if s is None: return None
    s = str(s).strip()
    return s if s else None

def norm_name(s):
    if not s: return None
    s = s.strip().strip('"').strip("'").lower()
    return re.sub(r'\s+', ' ', s)

# ── Load NZTCS names from Excel ────────────────────────────────────────────

print("Loading NZTCS names from Excel…")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Collect: canonical key → (original name, NZTCS entry summary)
name_to_key = {}   # original_name → norm_key (for reverse lookup)
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    name = clean(d.get('Current Species Name'))
    if name:
        name_to_key[name] = norm_name(name)

wb.close()
print(f"  {len(name_to_key)} unique species names")

# ── Load / seed cache ──────────────────────────────────────────────────────

try:
    with open(CACHE_FILE, encoding='utf-8') as f:
        cache = json.load(f)
    print(f"  Cache: {len(cache)} entries already resolved")
except FileNotFoundError:
    cache = {}

# ── GBIF species match ─────────────────────────────────────────────────────

def gbif_match(name):
    """Return GBIF match dict for name, or error dict."""
    url = ('https://api.gbif.org/v1/species/match?' +
           urlencode({'name': name, 'strict': 'false', 'verbose': 'false'}))
    for attempt in range(3):
        try:
            with urlopen(url, timeout=15) as resp:
                return json.loads(resp.read())
        except Exception as exc:
            if attempt == 2:
                return {'matchType': 'ERROR', '_error': str(exc)}
            time.sleep(0.5 * (attempt + 1))

to_query = [n for n in name_to_key if n not in cache]
print(f"  Querying GBIF for {len(to_query)} names "
      f"({len(name_to_key) - len(to_query)} cached)…")

done = errors = 0
t0 = time.time()

with ThreadPoolExecutor(max_workers=CONCURRENCY) as pool:
    futures = {pool.submit(gbif_match, name): name for name in to_query}
    for future in as_completed(futures):
        name = futures[future]
        result = future.result()
        cache[name] = result
        done += 1
        if result.get('matchType') == 'ERROR':
            errors += 1
        if done % 500 == 0 or done == len(to_query):
            elapsed = time.time() - t0
            rate = done / elapsed if elapsed > 0 else 0
            eta = (len(to_query) - done) / rate if rate > 0 else 0
            print(f"  {done}/{len(to_query)} — {rate:.0f}/s — ETA {eta:.0f}s — {errors} errors")
            with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(cache, f, ensure_ascii=False, separators=(',', ':'))

# Final cache save
with open(CACHE_FILE, 'w', encoding='utf-8') as f:
    json.dump(cache, f, ensure_ascii=False, separators=(',', ':'))

elapsed = time.time() - t0
print(f"  API queries done in {elapsed:.1f}s — {errors} errors")

# ── Match statistics ───────────────────────────────────────────────────────

match_types = Counter(v.get('matchType', 'NONE') for v in cache.values())
print("\nMatch type breakdown:")
for mt, count in match_types.most_common():
    print(f"  {mt}: {count}")

# ── Load existing nztcs.json ───────────────────────────────────────────────

print(f"\nLoading {NZTCS_JSON}…")
with open(NZTCS_JSON, encoding='utf-8') as f:
    nztcs = json.load(f)
print(f"  {len(nztcs)} existing entries")

# Remove old gbifkey entries so we start fresh
old_keys = [k for k in nztcs if k.startswith('gbifkey:')]
for k in old_keys:
    del nztcs[k]
print(f"  Removed {len(old_keys)} stale gbifkey entries")

# ── Build gbifkey entries ──────────────────────────────────────────────────
# For each GBIF match, add entries keyed by:
#   "gbifkey:{speciesKey}"   — species-level key (most important; occurrence
#                              records with any subsp. taxon key will still
#                              carry the species key in the speciesKey field,
#                              and we check both in the app lookup)
#   "gbifkey:{usageKey}"     — exact usage key for this name
#   "gbifkey:{acceptedKey}"  — accepted name key (handles synonyms)

added = skipped_low_conf = skipped_no_key = 0

# Group entries by speciesKey to detect conflicts before writing
from collections import defaultdict
species_key_entries = defaultdict(list)   # speciesKey → [(entry, name), ...]
usage_key_entries   = {}                   # usageKey → entry (always safe — subspecies-specific)

for name, result in cache.items():
    mt = result.get('matchType', 'NONE')
    if mt not in MATCH_TYPES:
        continue
    conf = result.get('confidence', 0)
    if conf < CONFIDENCE:
        skipped_low_conf += 1
        continue

    norm = name_to_key.get(name)
    if not norm or norm not in nztcs:
        skipped_no_key += 1
        continue

    entry = nztcs[norm]
    sk = result.get('speciesKey')
    uk = result.get('usageKey')

    if sk:
        species_key_entries[int(sk)].append(entry)
    if uk and uk != sk:
        # usageKey is specific to this exact name (subspecies key) — always safe
        usage_key_entries[int(uk)] = entry

# Add species keys only when all entries sharing that key agree on status
species_added = species_skipped = 0
for sk, entries in species_key_entries.items():
    lookup_key = 'gbifkey:' + str(sk)
    if lookup_key in nztcs:
        continue
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        nztcs[lookup_key] = entries[0]
        species_added += 1
        added += 1
    else:
        species_skipped += 1   # mixed subspecies — no species-level badge

# Always add usage keys (subspecies-specific, unambiguous)
usage_added = 0
for uk, entry in usage_key_entries.items():
    lookup_key = 'gbifkey:' + str(uk)
    if lookup_key not in nztcs:
        nztcs[lookup_key] = entry
        usage_added += 1
        added += 1

print(f"  Species keys added: {species_added} (skipped {species_skipped} with mixed subspecies)")
print(f"  Usage keys added: {usage_added}")

print(f"\nNew gbifkey entries added: {added}")
print(f"Skipped (low confidence): {skipped_low_conf}")
print(f"Skipped (name not in nztcs): {skipped_no_key}")
print(f"Total entries: {len(nztcs)}")

# ── Write updated nztcs.json ───────────────────────────────────────────────

import os
with open(NZTCS_JSON, 'w', encoding='utf-8') as f:
    json.dump(nztcs, f, ensure_ascii=False, separators=(',', ':'))
size_kb = os.path.getsize(NZTCS_JSON) / 1024
print(f"\nWritten: {NZTCS_JSON} ({size_kb:.1f} KB)")

# ── Spot checks ────────────────────────────────────────────────────────────
print("\nSpot checks (key lookup):")
checks = {
    2474403: "Zapornia pusilla (Baillon's crake) — speciesKey",
    7268804: "Zapornia pusilla affinis — usageKey (if exists)",
    7518721: "Egretta sacra — speciesKey",
    2481815: "Porzana tabuensis / Spotless crake",
}
for gbif_id, label in checks.items():
    k = 'gbifkey:' + str(gbif_id)
    v = nztcs.get(k)
    status = v['status'] if v else 'NOT FOUND'
    print(f"  {gbif_id} ({label}): {status}")
