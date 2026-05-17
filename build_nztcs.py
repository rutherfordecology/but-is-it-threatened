"""
Build nztcs.json from NZTCS Exported Data.xlsx

Produces a flat lookup keyed by lowercase scientific name.
Each entry uses the most-recently-assessed record (by Year Assessed).
Aliases (Previous Name, Alternative Names) also get index entries.

Output: nztcs.json (same directory as this script)
"""

import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

XLSX = r'C:\Users\User\Downloads\NZTCS Exported Data.xlsx'
OUT  = r'C:\Users\User\GBIF-Record-Finder\.claude\worktrees\infallible-babbage-f27395\nztcs.json'

def clean(s):
    """Return stripped string or None."""
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None

def norm_name(s):
    """
    Normalise a scientific name for lookup:
    - lowercase
    - strip leading/trailing quotes and whitespace
    - collapse internal whitespace
    """
    if not s:
        return None
    s = s.strip().strip('"').strip("'").lower()
    s = re.sub(r'\s+', ' ', s)
    # Strip authority — keep only genus + specific epithet (first 2 tokens)
    # BUT only if it looks like a binomial (2+ words). Uninomials stay as-is.
    tokens = s.split()
    if len(tokens) >= 2:
        # Heuristic: authority starts at first token that's not a lowercase word
        # Actually let's keep up to 3 tokens (subsp.) — GBIF names are usually 2
        # For robustness, just keep all tokens — we want to match exactly what GBIF returns
        pass
    return s

def split_alts(s):
    """Split Alternative Names field (comma or semicolon separated)."""
    if not s:
        return []
    parts = re.split(r'[;,]', str(s))
    return [p.strip() for p in parts if p.strip()]

print("Loading workbook…")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"  Columns: {len(headers)}")

# Main lookup: key → entry dict
# We keep the most recent assessment per species name
lookup = {}   # lowercase name → entry
aliases = {}  # lowercase alias → canonical lowercase name

rows_read = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    rows_read += 1
    d = dict(zip(headers, row))

    current_name = clean(d.get('Current Species Name'))
    if not current_name:
        continue

    year = d.get('Year Assessed')
    try:
        year = int(year) if year else None
    except (ValueError, TypeError):
        year = None

    entry = {
        'name':     current_name,
        'common':   clean(d.get('Preferred Common Name')),
        'maori':    clean(d.get('Preferred Māori Name')),
        'status':   clean(d.get('Status')),
        'category': clean(d.get('Category')),
        'year':     year,
        'bioStatus': clean(d.get('Bio Status')),
        'order':    clean(d.get('Order')),
        'family':   clean(d.get('Family')),
    }

    key = norm_name(current_name)
    if key:
        # Keep newer assessment if duplicate
        existing = lookup.get(key)
        if not existing or (year and (not existing['year'] or year > existing['year'])):
            lookup[key] = entry

        # Previous name → alias
        prev = clean(d.get('Previous Name'))
        if prev:
            pk = norm_name(prev)
            if pk and pk != key:
                aliases[pk] = key

        # Alternative names → aliases
        for alt in split_alts(d.get('Alternative Names')):
            ak = norm_name(alt)
            if ak and ak != key:
                aliases[ak] = key

wb.close()
print(f"  Rows read: {rows_read}")
print(f"  Unique names: {len(lookup)}")
print(f"  Aliases: {len(aliases)}")

# Merge aliases that don't shadow a real name
added = 0
for alias_key, canonical_key in aliases.items():
    if alias_key not in lookup and canonical_key in lookup:
        lookup[alias_key] = lookup[canonical_key]
        added += 1

print(f"  Alias entries added: {added}")

# ── Subspecies binomial + split-species + old-name aliases ────────────────
# Rule: only add a binomial alias when ALL subspecies that share that
# binomial have the SAME status.  Mixed-status subspecies (e.g. Tui with
# a Not Threatened mainland form and a NV Chatham Island form) must not
# generate an alias — showing the most-threatened one is misleading.

def is_latin_name(key, entry):
    """True if key uses only lowercase alpha/hyphen tokens and looks Latin."""
    parts = key.split()
    if len(parts) < 2:
        return False
    if not all(re.match(r'^[a-z][a-z\-]*$', p) for p in parts):
        return False
    return entry.get('name', '')[:1].isupper()

from collections import defaultdict

# --- 1. Subspecies binomial aliases (current 3-word names → 2-word binomial) ---
binomial_candidates = defaultdict(list)
for key, entry in lookup.items():
    parts = key.split()
    if len(parts) == 3 and is_latin_name(key, entry):
        binomial_candidates[parts[0] + ' ' + parts[1]].append(entry)

subsp_added = subsp_skipped = 0
for binomial, entries in binomial_candidates.items():
    if binomial in lookup:
        continue
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:               # all subspecies agree → safe alias
        lookup[binomial] = entries[0]
        subsp_added += 1
    else:
        subsp_skipped += 1               # mixed → no alias, no misleading badge

print(f"  Subspecies binomial aliases added: {subsp_added} (skipped {subsp_skipped} mixed-status)")

# --- 2. Old-name binomial aliases (4+ token alias keys → 2-word binomial) ---
# e.g. "porzana tabuensis tabuensis gmelin, 1789" → "porzana tabuensis"
# Collect candidates first, then apply the same all-agree rule.
old_binomial_candidates = defaultdict(list)
for key in list(lookup.keys()):
    parts = key.split()
    if len(parts) < 3:
        continue
    if not is_latin_name(key, lookup[key]):
        continue
    binomial = parts[0] + ' ' + parts[1]
    if binomial not in lookup:
        old_binomial_candidates[binomial].append(lookup[key])

old_binomial_added = old_binomial_skipped = 0
for binomial, entries in old_binomial_candidates.items():
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        lookup[binomial] = entries[0]
        old_binomial_added += 1
    else:
        old_binomial_skipped += 1

print(f"  Old-name binomial aliases added: {old_binomial_added} (skipped {old_binomial_skipped} mixed-status)")

# --- 3. Split-species aliases (genus + subspecies epithet, e.g. "zapornia affinis") ---
# When a subspecies is treated as a full species by some authorities.
# Only safe when the subspecies epithet maps to a unique status.
split_candidates = defaultdict(list)
for key in list(lookup.keys()):
    parts = key.split()
    if len(parts) != 3:
        continue
    if not is_latin_name(key, lookup[key]):
        continue
    genus, species_ep, subsp_ep = parts
    if subsp_ep == species_ep:
        continue  # nominate subspecies already covered by binomial alias
    split_key = genus + ' ' + subsp_ep
    if split_key not in lookup:
        split_candidates[split_key].append(lookup[key])

split_added = split_skipped = 0
for split_key, entries in split_candidates.items():
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        lookup[split_key] = entries[0]
        split_added += 1
    else:
        split_skipped += 1

print(f"  Split-species aliases added: {split_added} (skipped {split_skipped} mixed-status)")
print(f"  Total lookup entries: {len(lookup)}")

# Status summary
from collections import Counter
status_counts = Counter(v['status'] for v in lookup.values())
print("\nStatus distribution (unique entries):")
for status, count in status_counts.most_common():
    print(f"  {status}: {count}")

# Write JSON
print(f"\nWriting {OUT}…")
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(lookup, f, ensure_ascii=False, separators=(',', ':'))

import os
size_kb = os.path.getsize(OUT) / 1024
print(f"Done. File size: {size_kb:.1f} KB")
